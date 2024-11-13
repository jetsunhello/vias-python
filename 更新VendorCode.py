import os
from openpyxl import load_workbook


def update_b8_cell(directory, new_value="100100"):
    # 遍历指定目录下所有的 .xlsm 文件
    for filename in os.listdir(directory):
        if filename.endswith('.xlsm'):
            file_path = os.path.join(directory, filename)
            try:
                # 加载工作簿
                workbook = load_workbook(filename=file_path, keep_vba=True)

                # 对于工作簿中的每个工作表
                for sheetname in workbook.sheetnames:
                    worksheet = workbook[sheetname]
                    # 设置 B8 单元格的值
                    worksheet['B8'] = new_value

                # 保存更改后的工作簿
                workbook.save(file_path)
                print(f"Updated {file_path}")
            except Exception as e:
                print(f"Error processing {file_path}: {e}")


# 使用你的目录路径替换 'your_directory_path'
directory_path = 'YMTC ECOA'
update_b8_cell(directory_path)