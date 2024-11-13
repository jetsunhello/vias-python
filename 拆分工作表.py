import os
from openpyxl import load_workbook, Workbook


def split_sheets_to_files(directory):
    # 遍历指定目录下所有的 .xlsm 文件
    for filename in os.listdir(directory):
        if filename.endswith('.xlsm'):
            file_path = os.path.join(directory, filename)
            try:
                # 加载工作簿
                workbook = load_workbook(filename=file_path, keep_vba=False)

                # 创建以原文件名命名的新文件夹
                folder_name = os.path.splitext(filename)[0]
                output_folder = os.path.join(directory, folder_name)
                os.makedirs(output_folder, exist_ok=True)

                # 对于工作簿中的每个工作表
                for sheetname in workbook.sheetnames:
                    worksheet = workbook[sheetname]

                    # 创建新的工作簿并添加工作表
                    new_workbook = Workbook()
                    new_worksheet = new_workbook.active
                    new_worksheet.title = sheetname

                    # 复制数据
                    for row in worksheet.iter_rows(values_only=True):
                        new_worksheet.append(row)

                    # 保存新文件
                    new_file_path = os.path.join(output_folder, f"{sheetname}.xlsx")
                    new_workbook.save(new_file_path)
                    print(f"Created {new_file_path}")

            except Exception as e:
                print(f"Error processing {file_path}: {e}")


# 使用你的目录路径替换 'your_directory_path'
directory_path = 'YMTC ECOA'
split_sheets_to_files(directory_path)