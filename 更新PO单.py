import os
from openpyxl import load_workbook

# 定义料号和PO号的映射关系
material_po_mapping = {
    'A4000001': '4500022598',
    'A4000003': '4500022598',
    'A4000005': '4500022598',
    'A4000010': '4500022598',
    'A4000036': '4500022598',
    'A4000051': '4500022598',
    'A4000002': '4500023844',
    'A4000004': '4500023844',
    'A4000039': '4500023844',
    'A4000049': '4500023844',
    'A4000096': '4500027152',
    'A4000037': '4500028929',
    'A4000038': '4500028929',
    'A4000093': '4500028929',
    'A4000090': '4500028929'
}


def update_b2_cell(directory):
    # 遍历指定目录下所有的 .xlsm 文件
    for filename in os.listdir(directory):
        if filename.endswith('.xlsm'):
            file_path = os.path.join(directory, filename)
            try:
                # 加载工作簿
                workbook = load_workbook(filename=file_path, keep_vba=True)

                # 对于工作簿中的每个工作表
                for sheetname in workbook.sheetnames:
                    worksheet = workbook[sheetname]
                    # 获取 B6 单元格的值
                    b6_value = worksheet['B6'].value

                    # 检查 B6 单元格的值是否在映射关系中
                    if b6_value in material_po_mapping:
                        # 更新 B2 单元格的值
                        worksheet['B2'] = material_po_mapping[b6_value]

                # 保存更改后的工作簿
                workbook.save(file_path)
                print(f"Updated {file_path}")
            except Exception as e:
                print(f"Error processing {file_path}: {e}")


# 使用你的目录路径替换 'your_directory_path'
directory_path = 'YMTC ECOA'
update_b2_cell(directory_path)